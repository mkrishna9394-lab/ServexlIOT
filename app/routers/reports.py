from fastapi import APIRouter, Request, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.services.event_logger import log_event
from app.core.database import get_db
from app.core.templates import templates
from app.core.deps import require_user
from app.models import HistoricalValue, ConfiguredTag, ConfiguredMeter, Gateway, Site

router = APIRouter(prefix="/reports")


def is_super_admin(user):
    return user.role and user.role.name == "super_admin"


def get_customer_gateway_ids(db, user):
    if is_super_admin(user):
        return [g.id for g in db.query(Gateway).all()]

    site_ids = [
        s.id for s in db.query(Site)
        .filter(Site.customer_id == user.customer_id)
        .all()
    ]

    if not site_ids:
        return []

    return [
        g.id for g in db.query(Gateway)
        .filter(Gateway.site_id.in_(site_ids))
        .all()
    ]


def parse_dt(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%dT%H:%M")
    except Exception:
        return None


def clean_id(value):
    if value is None:
        return None

    value = str(value).strip()

    if value == "":
        return None

    return int(value)


def get_range_datetimes(range_type, start_datetime, end_datetime):
    now = datetime.now()

    if range_type == "1h":
        return now - timedelta(hours=1), now

    if range_type == "24h":
        return now - timedelta(hours=24), now

    if range_type == "7d":
        return now - timedelta(days=7), now

    if range_type == "30d":
        return now - timedelta(days=30), now

    return parse_dt(start_datetime), parse_dt(end_datetime)


def build_report_query(
    db,
    user,
    range_type=None,
    start_datetime=None,
    end_datetime=None,
    configured_tag_id=None,
    configured_meter_id=None,
    gateway_id=None
):
    q = (
        db.query(HistoricalValue, ConfiguredTag, ConfiguredMeter, Gateway)
        .join(ConfiguredTag, HistoricalValue.configured_tag_id == ConfiguredTag.id)
        .join(ConfiguredMeter, ConfiguredTag.configured_meter_id == ConfiguredMeter.id)
        .join(Gateway, ConfiguredMeter.gateway_id == Gateway.id)
    )

    allowed_gateway_ids = get_customer_gateway_ids(db, user)

    if not is_super_admin(user):
        if not allowed_gateway_ids:
            q = q.filter(False)
        else:
            q = q.filter(Gateway.id.in_(allowed_gateway_ids))

    start_dt, end_dt = get_range_datetimes(
        range_type,
        start_datetime,
        end_datetime
    )

    if start_dt:
        q = q.filter(HistoricalValue.timestamp >= start_dt)

    if end_dt:
        q = q.filter(HistoricalValue.timestamp <= end_dt)

    gateway_id = clean_id(gateway_id)
    configured_meter_id = clean_id(configured_meter_id)
    configured_tag_id = clean_id(configured_tag_id)

    if gateway_id:
        if is_super_admin(user) or gateway_id in allowed_gateway_ids:
            q = q.filter(Gateway.id == gateway_id)
        else:
            q = q.filter(False)

    if configured_meter_id:
        q = q.filter(ConfiguredMeter.id == configured_meter_id)

    if configured_tag_id:
        q = q.filter(ConfiguredTag.id == configured_tag_id)

    return q.order_by(HistoricalValue.timestamp.desc())


@router.get("")
def index(
    request: Request,
    run: str | None = Query(None),
    page: int = Query(1),
    range_type: str | None = Query(None),
    start_datetime: str | None = Query(None),
    end_datetime: str | None = Query(None),
    gateway_id: str | None = Query(None),
    configured_meter_id: str | None = Query(None),
    configured_tag_id: str | None = Query(None),
    db: Session = Depends(get_db),
    user=Depends(require_user)
):
    rows = []
    total_pages = 0
    PAGE_SIZE = 15

    gateway_ids = get_customer_gateway_ids(db, user)

    gateways = (
        db.query(Gateway)
        .filter(Gateway.id.in_(gateway_ids))
        .all()
    )

    configured_meters = (
        db.query(ConfiguredMeter)
        .filter(ConfiguredMeter.gateway_id.in_(gateway_ids))
        .all()
    )

    configured_tags = (
        db.query(ConfiguredTag)
        .join(ConfiguredMeter, ConfiguredTag.configured_meter_id == ConfiguredMeter.id)
        .filter(ConfiguredTag.is_active == True)
        .filter(ConfiguredMeter.gateway_id.in_(gateway_ids))
        .all()
    )

    if run == "1":
        query = build_report_query(
            db,
            user,
            range_type,
            start_datetime,
            end_datetime,
            configured_tag_id,
            configured_meter_id,
            gateway_id
        )

        total_records = query.count()

        rows = (
            query
            .offset((page - 1) * PAGE_SIZE)
            .limit(PAGE_SIZE)
            .all()
        )

        total_pages = (total_records + PAGE_SIZE - 1) // PAGE_SIZE

    return templates.TemplateResponse(
        "reports.html",
        {
            "request": request,
            "user": user,
            "page": page,
            "total_pages": total_pages,
            "gateways": gateways,
            "configured_meters": configured_meters,
            "configured_tags": configured_tags,
            "rows": rows,
            "show_report": run == "1",
            "filters": {
                "range_type": range_type or "",
                "start_datetime": start_datetime or "",
                "end_datetime": end_datetime or "",
                "gateway_id": gateway_id or "",
                "configured_meter_id": configured_meter_id or "",
                "configured_tag_id": configured_tag_id or "",
            },
        },
    )


@router.get("/export.xlsx")
def export(
    range_type: str | None = Query(None),
    start_datetime: str | None = Query(None),
    end_datetime: str | None = Query(None),
    gateway_id: str | None = Query(None),
    configured_meter_id: str | None = Query(None),
    configured_tag_id: str | None = Query(None),
    db: Session = Depends(get_db),
    user=Depends(require_user)
):
    rows = (
        build_report_query(
            db,
            user,
            range_type,
            start_datetime,
            end_datetime,
            configured_tag_id,
            configured_meter_id,
            gateway_id
        )
        .limit(50000)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Historical Data"

    headers = [
        "Timestamp",
        "Gateway",
        "Meter",
        "Tag",
        "Value",
        "Unit",
        "Source"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="0B5E7A")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for hv, tag, meter, gateway in rows:
        ws.append([
            hv.timestamp.strftime("%Y-%m-%d %H:%M:%S") if hv.timestamp else "",
            gateway.code,
            meter.name,
            tag.display_name,
            hv.value,
            tag.unit,
            hv.source,
        ])

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = max_len + 3

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    
    log_event(db, user, "Reports", "Export Excel", "Historical report exported")

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=iiot_historical_report.xlsx"
        }
    )